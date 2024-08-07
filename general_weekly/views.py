import os
from datetime import datetime, timedelta

from django.apps import apps
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.http import FileResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

from commondata.forms import DateForm, DateSelectionForm
from commondata.models import Department
from .decorators import (general_weekly_check_user_department, check_general_weekly_summary_report_creator,
                         general_weekly_check_user)
from .forms import WeeklyReportForm
from .models import WeeklyUserDepartment, WeeklyReport, WeeklyCreatorsSummaryReport
from .utils import friday_of_week


@login_required
@general_weekly_check_user
def general_weekly(request):
    """
    Список отчётов
    """
    # Получаем текущего пользователя
    user = request.user
    # Получаем филиалы, к которым пользователь имеет отношение
    user_departments = WeeklyUserDepartment.objects.filter(user=user).values_list('department', flat=True)
    # Получаем все отчёты, связанные с этими филиалами, и сортируем их по дате отчёта
    reports = WeeklyReport.objects.filter(department__in=user_departments).order_by('-report_date')
    # Если форма отправлена методом POST
    if request.method == 'POST':
        form = DateForm(request.POST)
        if form.is_valid():
            selected_date = form.cleaned_data['selected_date']
            # Фильтруем отчёты по выбранной дате
            reports = reports.filter(report_date=friday_of_week(selected_date))
    else:
        form = DateForm(initial={'selected_date': friday_of_week(datetime.now())})
    for report in reports:
        report.user_full_name = f"{report.author.last_name} {report.author.first_name}"
    first_summary_report = WeeklyCreatorsSummaryReport.objects.first()
    if first_summary_report:
        summary_reports_creators = [user.username for user in first_summary_report.creators.all()]
    return render(request, 'general_weekly/report_list.html',
                  {'reports': reports,
                   'form': form,
                   'summary_reports_creators': summary_reports_creators})


@login_required
@general_weekly_check_user_department
def general_weekly_report_details(request, report_id):
    """
    Детализированный отчёт
    :param request:
    :param report_id:
    :return:
    """
    # Получаем объект отчёта по его id
    report = get_object_or_404(WeeklyReport, pk=report_id)
    # Объединяем имя и фамилию пользователя через пробел
    user_full_name = f"{report.author.last_name} {report.author.first_name}"
    return render(request, 'general_weekly/report_details.html', {'report': report, 'user_full_name': user_full_name})


@login_required
def add_general_weekly_report(request):
    """
    Добавление отчёта
    :param request:
    :return:
    """
    if request.method == 'POST':
        # Создаем экземпляр формы ежедневного отчёта, передавая текущего пользователя и данные из POST-запроса
        form = WeeklyReportForm(request.user, request.POST)  # Передаем объект пользователя в форму
        if form.is_valid():  # Проверяем валидность данных формы
            weekly_report = form.save(commit=False)  # Создаем объект ежедневного отчёта, не сохраняя его в базу данных
            weekly_report.author = request.user  # Устанавливаем автора отчёта
            # Проверяем наличие отчёта для выбранного филиала и даты
            if (WeeklyReport.objects.filter(department=weekly_report.department)
                    .filter(report_date=weekly_report.report_date).exists()):
                return render(request, 'general_weekly/denied_add_report.html',
                              {'department': weekly_report.department.name,
                               'report_date': weekly_report.report_date.strftime('%d.%m.%Y')})
            weekly_report.save()  # Сохраняем отчёт в базу данных
            # Перенаправляем пользователя на success_page в случае успешного сохранения отчёта
            return redirect(success_page)
    else:
        # Получаем первое подразделение пользователя, если оно есть
        first_user_department = WeeklyUserDepartment.objects.filter(user=request.user).first()
        if first_user_department:
            first_department = first_user_department.department.first()
        else:
            # Иначе получаем первое подразделение в системе
            first_department = Department.objects.first()
        # Создаем экземпляр формы ежедневного отчёта с начальными данными,
        # передавая текущего пользователя и первое доступное пользователю подразделение
        form = WeeklyReportForm(request.user, initial={
            'department': first_department})  # Передаем объект пользователя и начальные значения в форму
    # Отображаем шаблон add_general_weekly_report.html с переданной формой
    return render(request, 'general_weekly/add_general_weekly_report.html', {'form': form})


@login_required
@general_weekly_check_user_department
def edit_general_weekly_report(request, report_id):
    """
    Редактирование отчёта
    """
    # Получаем отчёт по его идентификатору
    report = get_object_or_404(WeeklyReport, id=report_id)
    # Получаем текущее время в часовом поясе Москвы
    current_time = timezone.now()
    # Проверяем, прошло ли меньше 1 часа с момента создания отчёта
    if current_time - report.created_at > timedelta(hours=1):
        # Если прошло более часа, переходим на страницу с сообщением об ошибке
        return render(request, 'general_weekly/error_edit_report.html')
    if request.method == 'POST':
        # Если запрос метода POST, обрабатываем форму
        form = WeeklyReportForm(request.user, request.POST, instance=report)
        if form.is_valid():
            # Если форма валидна, то сохраняем отчёт
            weekly_report = form.save(commit=False)
            weekly_report.author = request.user
            weekly_report.save()
            # Перенаправляем пользователя на страницу успешного завершения
            return redirect(success_page)
    else:
        # Если запрос метода GET, отображаем форму для редактирования
        form = WeeklyReportForm(request.user, instance=report)
    # Возвращаем HTML-страницу с формой для редактирования отчёта
    return render(request, 'general_weekly/edit_general_weekly_report.html', {'form': form, 'report_id': report_id})


@login_required
@check_general_weekly_summary_report_creator
def general_weekly_summary_report(request):
    """
    Сводный отчёт
    """

    def get_users_for_department(department: str) -> list:
        """
        department: строковое название филиала
        users: модель, хранящая отношения пользователя и филиалы
        return: список пользователей для данного филиала
        """
        # return [user for user in users if department in user.department]
        users = WeeklyUserDepartment.objects.filter(department__name=department)
        # Получаем всех создателей сводного отчета
        creators = WeeklyCreatorsSummaryReport.objects.values_list('creators', flat=True)
        # Исключаем создателей из списка пользователей
        users = users.exclude(user__in=creators)
        # Извлекаем всех пользователей из найденных записей
        users = [f'{user.user.last_name} {user.user.first_name}' for user in users]
        return users

    if request.method == 'POST':
        # Если запрос метода POST, обрабатываем форму
        form = DateSelectionForm(request.POST)
        if form.is_valid():
            # Если форма действительна, извлекаем выбранную дату из формы
            date = friday_of_week(form.cleaned_data['report_date'])
    else:
        form = DateSelectionForm(initial={'report_date': friday_of_week(datetime.now())})
        date = friday_of_week(datetime.now())
    reports = WeeklyReport.objects.filter(report_date=date)
    # Получаем список подразделений и пользователей, без отчётов за дату date
    departments_without_reports = [': '.join([department.name, ', '.join(get_users_for_department(department.name))])
                                   for department in Department.objects.all().exclude(weeklyreport__report_date=date)
                                   .order_by('name')]
    return render(request, 'general_weekly/summary_report.html',
                  {'form': form, 'reports': reports,
                   'departments_without_reports': departments_without_reports})


@login_required
@check_general_weekly_summary_report_creator
def generate_general_weekly_summary_report(request):
    """
    Генерация сводного еженедельного отчёта
    """
    if request.method == 'POST':
        # Получаем данные формы
        form = DateSelectionForm(request.POST)
        if form.is_valid():
            # Извлекаем выбранную дату из формы
            selected_date = friday_of_week(form.cleaned_data['report_date'])
            # Получаем экземпляр класса конфигурации приложения general_weekly
            general_weekly_config = apps.get_app_config('general_weekly')
            # Получаем путь к папке для сохранения отчётов
            path_to_reports = general_weekly_config.PATH_TO_SAVE
            report_name = create_general_weekly_report(selected_date, path_to_reports)
            return FileResponse(open(report_name, 'rb'), as_attachment=True,
                                filename=report_name)


def general_weekly_access_denied_page(request):
    return render(request, 'general_weekly/access_denied_page.html')


@login_required
def success_page(request):
    return render(request, 'general_weekly/success_page.html')


def create_general_weekly_report(selected_date, path_to_reports):
    """
    Генерирует сводный ежедневный отчёт за указанную дату и возвращает путь к файлу отчёта.

    :param selected_date: Дата отчета, тип datetime.date
    :param path_to_reports: Путь к папке для сохранения отчётов
    :return: Путь к сгенерированному файлу отчёта
    """
    # описание структуры готового отчёта
    report_structure = (
        # 1 значение - строка, "Номер подпункта"
        # 2 значение - строка, "Название раздела"
        # 3 тип - булевая, True - суммируем числовые значения, False - формируем строковое описание
        ('1', 'Экономический эффект (сумма, млн. руб без НДС)', True),
        ('1.1', 'Наиболее значимый пример:', False),
        ('2', 'Выявлено неучтенных ТМЦ (сумма, млн. руб без НДС)', True),
        ('2.1', 'Наиболее значимый пример:', False),
        ('3', 'Входной контроль (сумма, млн. руб без НДС)', True),
        ('3.1', 'Наиболее значимый пример:', False),
        ('4', 'Количество запросов, актов реагирования от контрольно-надзорных и правоохранительных органов, '
              'поступивших в отчетном периоде', True),
        ('4.1', 'Наиболее значимый пример:', False),
        ('5', 'Количество запросов, актов реагирования, поручений поступивших из территориальных органов '
              'власти, поступивших в отчетном периоде', True),
        ('5.1', 'Наиболее значимый пример:', False),
        ('6', 'Направлено заявлений в правоохранительные органы для защиты интересов Компании', True),
        ('6.1', 'Наиболее значимый пример:', False),
        ('7', 'Проведено встреч, рабочих групп с сотрудниками правоохранительных, контрольно-надзорными органов', True),
        ('7.1', 'Наиболее значимый пример:', False),
        ('8', 'Выявлено фактов антикорпоративных проявлений', True),
        ('8.1', 'Наиболее значимый пример:', False),
        ('9', 'Инициировано служебных проверок', True),
        ('9.1', 'Наиболее значимый пример:', False),
        ('10', 'Значимая информация, факты, события, риски и т.д.', False)
    )

    def get_last_saturday_and_current_friday_dates(current_date):
        # Определить день недели (0 - понедельник, 6 - воскресенье)
        current_weekday = current_date.weekday()
        # Вычислить дату прошлой субботы
        _last_saturday = current_date - timedelta(days=current_weekday + 2)
        # Вычислить дату текущей пятницы
        _current_friday = current_date + timedelta(days=(4 - current_weekday))
        # Преобразовать даты в строки в формате '%d.%m.%Y'
        last_saturday_str = _last_saturday.strftime('%d.%m.%Y')
        current_friday_str = _current_friday.strftime('%d.%m.%Y')
        # Вернуть кортеж из строк с датами
        return last_saturday_str, current_friday_str

    # Префикс названия отчёта
    prefix_report_name = 'Еженедельный отчёт филиалов'
    # Формируем имя файла отчёта на основе выбранной даты
    report_name = os.path.join(path_to_reports, f'{prefix_report_name} за '
                                                f'{selected_date.strftime("%d.%m.%Y")}.xlsx')
    # Создаём xlsx-файл отчёта
    wb = Workbook()
    sheet = wb.active
    sheet.title = f'Отчёт {selected_date.strftime("%d.%m.%Y")}'
    # Заполняем отчёт
    # Масштаб 55%
    sheet.page_setup.scale = 55  # НЕ РАБОТАЕТ!!!
    # Границы ячейки
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    # Форматируем 1 строку
    sheet.merge_cells(f'A1:C1')
    sheet.column_dimensions['A'].width = 4  # 15
    sheet.column_dimensions['B'].width = 44  # 59
    sheet.column_dimensions['C'].width = 115  # 247
    sheet.row_dimensions[1].height = 43  # 100
    last_saturday, current_friday = get_last_saturday_and_current_friday_dates(selected_date)
    sheet['A1'].value = f'Еженедельный отчёт\n' \
                        f'эффективности работы СБ филиалов\n' \
                        f'c {last_saturday} г. по {current_friday} г.'
    sheet['A1'].font = Font(name='Tahoma', bold=True, size=11)  # size=24)
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet['A1'].border = border
    sheet['B1'].border = border
    sheet['C1'].border = border
    # Получаем все отчёты за выбранную дату, отсортированные по названию филиала
    reports = WeeklyReport.objects.filter(report_date=selected_date).order_by('department__name')
    current_row = 2
    start_row = 0
    for num, paragraph in enumerate(report_structure, start=1):
        if paragraph[2]:
            # суммируем значения
            sheet[f'A{current_row}'] = paragraph[0]
            sheet[f'B{current_row}'] = paragraph[1]
            sheet[f'C{current_row}'] = reports.aggregate(total_field_sum=Sum(f'field{num}'))[f'total_field_sum']
            sheet[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            sheet[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            sheet[f'C{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            sheet[f'A{current_row}'].font = Font(name='Tahoma', size=11)  # size=24)
            sheet[f'B{current_row}'].font = Font(name='Tahoma', size=11)  # size=24)
            sheet[f'B{current_row}'].fill = PatternFill(fill_type='solid',
                                                        start_color='ffff00', end_color='ffff00')
            sheet[f'C{current_row}'].font = Font(name='Tahoma', size=11)  # size=24)
            sheet[f'A{current_row}'].border = border
            sheet[f'B{current_row}'].border = border
            sheet[f'C{current_row}'].border = border
            current_row += 1
        else:
            # формируем строковое описание
            start_row = current_row
            sheet[f'A{current_row}'] = paragraph[0]
            sheet[f'B{current_row}'] = paragraph[1]
            sheet[f'A{current_row}'].font = Font(name='Tahoma', size=11)  # size=24)
            sheet[f'B{current_row}'].font = Font(name='Tahoma', size=11)  # size=24)
            sheet[f'A{current_row}'].border = border
            sheet[f'B{current_row}'].border = border
            sheet[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            sheet[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            for report in reports:
                if getattr(report, f'field{num}').strip():
                    sheet.row_dimensions[current_row].auto_size = True
                    sheet[f'C{current_row}'] = CellRichText(
                        TextBlock(InlineFont(rFont='Tahoma', sz=11, b=True), report.department.name),
                        '     ',
                        TextBlock(InlineFont(rFont='Tahoma', sz=11), getattr(report, f'field{num}').strip()),
                    )
                    sheet[f'C{current_row}'].font = Font(name='Tahoma', size=11)  # size=24)
                    sheet[f'C{current_row}'].border = border
                    sheet[f'C{current_row}'].alignment = Alignment(horizontal='left', vertical='center',
                                                                   wrap_text=True)

                    current_row += 1
            if start_row < current_row:
                sheet.merge_cells(f'A{start_row}:A{current_row - 1}')
                sheet.merge_cells(f'B{start_row}:B{current_row - 1}')
    # закрасим ячейку B в посл. строке жёлтым
    sheet[f'B{start_row}'].fill = PatternFill(fill_type='solid',
                                              start_color='ffff00', end_color='ffff00')
    wb.save(report_name)
    return report_name


def get_users_for_department(department: str) -> list:
    """
    department: строковое название филиала
    users: модель, хранящая отношения пользователя и филиалы
    return: список пользователей для данного филиала
    """
    # return [user for user in users if department in user.department]
    users = WeeklyUserDepartment.objects.filter(department__name=department)
    # Получаем всех создателей сводного отчета
    creators = WeeklyCreatorsSummaryReport.objects.values_list('creators', flat=True)
    # Исключаем создателей из списка пользователей
    users = users.exclude(user__in=creators)
    # Извлекаем всех пользователей из найденных записей
    users = [f'{user.user.last_name} {user.user.first_name}' for user in users]
    return users
