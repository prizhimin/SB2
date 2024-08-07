import os
import socket
import shutil
from django.conf import settings
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.http import FileResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.apps import apps

from .forms import InvestigationForm, AttachedFileForm, DateForm
from .models import Investigation, AttachedFile, InvestigationUserDepartment, InvestigationCreatorsSummaryReport
from .decorators import check_user_department, check_summary_report_creator
from django.utils import timezone
from commondata.forms import DateRangeForm
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from datetime import date
from datetime import datetime


@login_required
def investigation_list(request):
    # Получаем текущего пользователя
    user = request.user
    # Получаем филиалы, к которым пользователь имеет отношение
    user_departments = InvestigationUserDepartment.objects.filter(user=user).values_list('department', flat=True)
    # Получаем СЗ, к которым у пользователя есть доступ
    investigations = Investigation.objects.filter(department__in=user_departments).order_by('-order_date')
    # Если форма была отправлена методом POST, обрабатываем её
    if request.method == 'POST':
        form = DateForm(request.POST)
        if form.is_valid():
            selected_date = form.cleaned_data['selected_date']
            # Фильтруем служебные проверки по выбранной дате
            investigations = investigations.filter(order_date=selected_date)
    else:
        # Если форма не отправлена, создаем форму с начальной датой для отчета
        form = DateForm(initial={'selected_date': timezone.now().astimezone(tz=timezone.get_current_timezone())})
    # Получаем список создателей сводных отчетов, если таковые имеются
    summary_reports_creators = []
    first_summary_report = InvestigationCreatorsSummaryReport.objects.first()
    if first_summary_report:
        # Сохраняем имена пользователей, создавших первый сводный отчет
        summary_reports_creators = [user.username for user in first_summary_report.creators.all()]
    paginator = Paginator(investigations, 14)  # X расследования на страницу
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'investigations/investigation_list.html',
                  {'page_obj': page_obj,
                   'form': form,
                   'summary_reports_creators': summary_reports_creators})


@login_required
def add_investigation(request):
    if request.method == 'POST':
        form = InvestigationForm(request.user, request.POST)
        if form.is_valid():
            form.save()
            return redirect('investigations:investigation_list')
    else:
        first_user_department = InvestigationUserDepartment.objects.filter(user=request.user).first()
        form = InvestigationForm(request.user, initial={'department': first_user_department})
    return render(request, 'investigations/add_investigation.html', {'form': form})


@login_required
@check_user_department
def investigation_detail(request, pk):
    investigation = get_object_or_404(Investigation, pk=pk)
    return render(request, 'investigations/investigation_detail.html',
                  {'investigation': investigation})


@login_required
@check_user_department
def edit_investigation(request, pk):
    # Получаем объект по первичному ключу или возвращаем 404 ошибку, если объект не найден
    investigation = get_object_or_404(Investigation, pk=pk)

    # Проверяем, что это POST запрос
    if request.method == "POST":
        # Передаем экземпляр объекта и данные формы
        form = InvestigationForm(request.user, request.POST, instance=investigation)
        if form.is_valid():  # Проверка на валидность данных формы
            investigation = form.save(commit=False)  # Сохраняем объект формы в переменную, но пока не записываем в БД
            investigation.save()  # Теперь сохраняем изменения в БД
            return redirect('investigations:investigation_detail',
                            pk=investigation.pk)  # Перенаправление на страницу детального просмотра
    else:
        form = InvestigationForm(request.user, instance=investigation)  # Создаем форму с начальными данными объекта

    # Отображаем страницу с формой
    return render(request, 'investigations/edit_investigation.html', {'form': form,
                                                                      'investigation_pk': pk,
                                                                      'has_attach': investigation.has_attach})


# Функции для работы с прикрепленными файлами

@login_required
# @check_user_department
def attach_file(request, investigation_id):
    investigation = get_object_or_404(Investigation, id=investigation_id)
    if request.method == 'POST':
        form = AttachedFileForm(request.POST, request.FILES)
        if form.is_valid():
            form.save(investigation=investigation)
            return redirect('investigations:manage_attach', pk=investigation.pk)
    else:
        form = AttachedFileForm()
    return render(request, 'investigations/attach_file_form.html', {'form': form, 'investigation': investigation})


@login_required
# @check_user_department
def delete_file(request, file_id):
    attached_file = get_object_or_404(AttachedFile, id=file_id)
    investigation_id = attached_file.investigation.id
    if request.method == 'POST':
        attached_file.delete()
        file_path = os.path.join(settings.MEDIA_ROOT, 'invests', str(investigation_id),
                                 os.path.basename(attached_file.file.name))
        if os.path.exists(file_path):
            os.remove(file_path)
        return redirect('investigations:manage_attach', pk=investigation_id)
    return render(request,
                  'investigations/delete_file_confirm.html',
                  {'attached_file': attached_file, 'short_name': os.path.basename(attached_file.file.name)})


@login_required
def download_file(request, file_id):
    attached_file = get_object_or_404(AttachedFile, id=file_id)
    file_path = os.path.join(settings.MEDIA_ROOT, 'invests', str(attached_file.investigation.id),
                             os.path.basename(attached_file.file.name))
    response = FileResponse(open(file_path, 'rb'), as_attachment=True,
                            filename=file_path)
    return response


@login_required
def delete_investigation(request, pk):
    investigation = get_object_or_404(Investigation, pk=pk)
    if request.method == 'POST':
        for attached_file in investigation.attached_files.all():
            file_path = os.path.join(settings.MEDIA_ROOT, 'invests', str(investigation.id),
                                     os.path.basename(attached_file.file.name))
            if os.path.exists(file_path):
                os.remove(file_path)
        investigation_dir = os.path.join(settings.MEDIA_ROOT, 'invests', str(investigation.id))
        if os.path.exists(investigation_dir):
            shutil.rmtree(investigation_dir)
        investigation.delete()
        return redirect('investigations:investigation_list')
    return render(request, 'investigations/investigation_confirm_delete.html',
                  {'investigation': investigation})



@login_required
def manage_attach(request, pk):
    investigation = get_object_or_404(Investigation, id=pk)
    attached_files = AttachedFile.objects.filter(investigation=investigation)
    return render(request, 'investigations/manage_attach.html',
                  {'investigation': investigation,
                   'attached_files': attached_files})


@login_required
def download_attaches_zip(request, investigation_id):
    investigation = get_object_or_404(Investigation, id=investigation_id)
    if investigation.has_attach() > 0:
        attached_files = AttachedFile.objects.filter(investigation=investigation)
        for attached_file in attached_files:
            print(os.path.join(settings.MEDIA_ROOT, str(attached_file.file)))
        return HttpResponse('ZIPZIPZIP!!!')
    else:
        return HttpResponse('Нечего скачивать')


@login_required
@check_summary_report_creator
def summary_report(request):
    investigations = None
    # Получаем список создателей сводных отчетов, если таковые имеются
    summary_reports_creators = []
    first_summary_report = InvestigationCreatorsSummaryReport.objects.first()
    if first_summary_report:
        # Сохраняем имена пользователей, создавших первый сводный отчет
        summary_reports_creators = [user.username for user in first_summary_report.creators.all()]
    if request.method == "POST":
        form = DateRangeForm(request.POST)
        if form.is_valid():
            # Обработка данных формы
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
            investigations = (Investigation.objects.filter(order_date__range=(start_date, end_date))
                              .order_by('-order_date'))
    else:
        # Получить текущий год
        current_year = date.today().year
        # Устанавливаем первоначальный диапазон дат, равный текущему году
        form = DateRangeForm(initial={'start_date': f'01.01.{current_year}',
                                      'end_date': f'31.12.{current_year}'})
        investigations = (Investigation.objects
                          .filter(order_date__range=(datetime.strptime(f'01.01.{current_year}', '%d.%m.%Y').date(),
                                                     datetime.strptime(f'31.12.{current_year}', '%d.%m.%Y').date()))
                          .order_by('-order_date'))
    return render(request, 'investigations/investigations_report.html',
                  {'investigations': investigations,
                   'form': form,
                   'summary_reports_creators': summary_reports_creators})


@login_required
@check_summary_report_creator
def generate_summary_report(request, operation_id):
    if request.method == "POST":
        form = DateRangeForm(request.POST)
        if form.is_valid():
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
            match operation_id:
                # Все организации
                case 0:
                    investigations = (Investigation.objects.filter(order_date__range=(start_date, end_date))
                                      .order_by('order_date'))
                # ПАО "Т Плюс"
                case 1:
                    investigations = (Investigation.objects.filter(order_date__range=(start_date, end_date))
                                      .filter(department__name__startswith='ПАО "Т Плюс"')
                                      .order_by('order_date'))
                # АО "ЭнергосбыТ Плюс"
                case 2:
                    investigations = (Investigation.objects.filter(order_date__range=(start_date, end_date))
                                      .filter(department__name__startswith='АО "ЭнергосбыТ Плюс"')
                                      .order_by('order_date'))
                # АО "ЭнергоремонТ Плюс"
                case 3:
                    investigations = (Investigation.objects.filter(order_date__range=(start_date, end_date))
                                      .filter(department__name__startswith='АО "ЭнергоремонТ Плюс"')
                                      .order_by('order_date'))
            # Формируем пустой xlsx-файл с шапкой
            wb = Workbook()
            ws = wb.active
            # Создаём "шапку"
            head_sheet = [  # (текст ячейки, ширина ячейки)
                ('№ п/п', 6), ('Юридическое лицо', 23), ('Филиал', 42),
                ('Дата приказа', 12), ('Номер приказа', 12), ('Тип служебной  проверки', 49),
                ('Краткая фабула проверки', 50), ('Инициатор проверки', 13), ('Дата окончания проверки', 15),
                ('Дата окончания при продлении', 15), ('Текущее состояние по проверке', 14), ('Ущерб (млн. руб.)', 8),
                ('Возмещено/предотвращено\n(млн. руб)', 31), ('Краткое описание итогов', 40),
                ('Количество работников, привлечённых\nк дисциплионарной ответственности\n(депремировано)', 26),
                ('Количество работников, привлечённых\nк дисциплионарной ответственности\n(уволено)', 26),
                ('Количество работников, привлечённых\nк дисциплионарной ответственности\n(понижено в должности)', 26),
                ('Количество работников, привлечённых\nк дисциплионарной ответственности\n(выговор)', 26),
                ('Количество работников, привлечённых\nк дисциплионарной ответственности\n(замечание)', 26),
                ('Ссылка на папку\nс материалами проверки', 27)
            ]
            #
            for idx, col in enumerate('ABCDEFGHIJKLMNOPQRST'):
                ws[f'{col}1'] = head_sheet[idx][0]
                ws[f'{col}1'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                ws[f'{col}1'].font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
                ws[f'{col}1'].fill = PatternFill(fill_type='solid', fgColor='70AD47')
                ws[f'{col}1'].border = Border(top=Side(border_style='thin'),
                                              left=Side(border_style='thin'),
                                              right=Side(border_style='thin'),
                                              bottom=Side(border_style='thin'))
                ws.column_dimensions[col].width = head_sheet[idx][1]
            ws.auto_filter.ref = 'A1:T1'
            # Заполняем таблицу данными о СЗ
            row_number = 2
            for investigation in investigations:
                # № п/п
                ws[f'A{row_number}'] = row_number-1
                ws[f'A{row_number}'].alignment = Alignment(horizontal='center')
                # ЮЛ и Филила
                ws[f'B{row_number}'] = investigation.department.name.split('-')[0]
                ws[f'C{row_number}'] = investigation.department.name.split('-')[1]
                # Дата приказа
                ws[f'D{row_number}'] = investigation.order_date
                ws[f'D{row_number}'].number_format = 'DD.MM.YYYY'
                ws[f'D{row_number}'].alignment = Alignment(horizontal='center')
                # Номер приказа
                ws[f'E{row_number}'] = investigation.order_num
                ws[f'E{row_number}'].alignment = Alignment(horizontal='center')
                # Тип служебной проверки
                ws[f'F{row_number}'] = investigation.get_inspection_type_display()
                # Краткая фабула проверки
                ws[f'G{row_number}'] = investigation.brief_summary
                ws[f'G{row_number}'].alignment = Alignment(wrap_text=True)
                # Инициатор проверки
                ws[f'H{row_number}'] = investigation.initiator
                ws[f'H{row_number}'].alignment = Alignment(horizontal='center')
                # Дата окончания проверки
                ws[f'I{row_number}'] = investigation.end_date
                ws[f'I{row_number}'].number_format = 'DD.MM.YYYY'
                ws[f'I{row_number}'].alignment = Alignment(horizontal='center')
                # Дата окончания при продлении
                ws[f'J{row_number}'] = investigation.extended_end_date
                ws[f'J{row_number}'].number_format = 'DD.MM.YYYY'
                ws[f'J{row_number}'].alignment = Alignment(horizontal='center')
                # Текущее состояние по проверке
                ws[f'K{row_number}'] = investigation.get_status_display()
                ws[f'K{row_number}'].alignment = Alignment(horizontal='center')
                # Ущерб (млн. руб.)
                ws[f'L{row_number}'] = investigation.damage_amount
                ws[f'L{row_number}'].number_format = '#,##0.000'
                ws[f'L{row_number}'].alignment = Alignment(horizontal='center')
                # Возмещено/предотвращено(млн. руб)
                ws[f'M{row_number}'] = investigation.recovered_amount
                ws[f'M{row_number}'].number_format = '#,##0.000'
                ws[f'M{row_number}'].alignment = Alignment(horizontal='center')
                # Краткое описание итого
                ws[f'N{row_number}'] = investigation.outcome_summary
                ws[f'N{row_number}'].alignment = Alignment(wrap_text=True)
                # Количество работников, привлечённых к дисциплионарной ответственности (депремировано)
                ws[f'O{row_number}'] = investigation.num_employees_discipline_demotion
                ws[f'O{row_number}'].number_format = '#,##0.000'
                ws[f'O{row_number}'].alignment = Alignment(horizontal='center')
                # Количество работников, привлечённых к дисциплинарной ответственности (уволено)
                ws[f'P{row_number}'] = investigation.num_employees_discipline_fired
                ws[f'P{row_number}'].number_format = '#,##0.000'
                ws[f'P{row_number}'].alignment = Alignment(horizontal='center')
                # Количество работников, привлечённых к дисциплинарной ответственности (понижено в должности)
                ws[f'Q{row_number}'] = investigation.num_employees_discipline_reduction
                ws[f'Q{row_number}'].number_format = '#,##0.000'
                ws[f'Q{row_number}'].alignment = Alignment(horizontal='center')
                # Количество работников, привлечённых к дисциплинарной ответственности (выговор)
                ws[f'R{row_number}'] = investigation.num_employees_discipline_reprimand
                ws[f'R{row_number}'].number_format = '#,##0.000'
                ws[f'R{row_number}'].alignment = Alignment(horizontal='center')
                # Количество работников, привлечённых к дисциплинарной ответственности (замечание)
                ws[f'S{row_number}'] = investigation.num_employees_discipline_warning
                ws[f'S{row_number}'].number_format = '#,##0.000'
                ws[f'S{row_number}'].alignment = Alignment(horizontal='center')
                # Ссылка на материалы проверки
                if investigation.has_attach():
                    ws[f'T{row_number}'].hyperlink = (f'http://{socket.gethostbyname(socket.gethostname())}'
                                                      f'/investigations/{investigation.id}/manage_attach')
                    ws[f'T{row_number}'].value = 'Ссылка'
                    ws[f'T{row_number}'].style = 'Hyperlink'
                    ws[f'T{row_number}'].alignment = Alignment(horizontal='center')
                # Раскрашиваем строку
                for col in 'ABCDEFGHIJKLMNOPQRST':
                    ws[f'{col}{row_number}'].border = Border(top=Side(border_style='thin'),
                                                             left=Side(border_style='thin'),
                                                             right=Side(border_style='thin'),
                                                             bottom=Side(border_style='thin'))
                    if not (row_number % 2):
                        ws[f'{col}{row_number}'].fill = PatternFill(fill_type='solid', fgColor='E2EFDA')
                # Переходим к следующей строчке
                row_number += 1
            # Получаем экземпляр класса конфигурации приложения daily
            daily_config = apps.get_app_config('investigations')
            path_to_reports = daily_config.PATH_TO_SAVE
            # Префикс названия отчёта
            prefix_report_name = 'Сводный реестр по служебным проверкам'
            report_name = os.path.join(path_to_reports, f'{prefix_report_name}.xlsx')
            wb.save(report_name)
            response = FileResponse(open(report_name, 'rb'), as_attachment=True,
                                    filename=report_name)
            return response
        else:
            return HttpResponse('Invalid form data')  # Сообщение о неверных данных формы
    else:
        return HttpResponse('Only POST requests are allowed')  # Сообщение о неправильном методе запроса
