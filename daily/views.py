from django.shortcuts import render, get_object_or_404
from datetime import timedelta, datetime
from django.contrib.auth.decorators import login_required
from django.http import FileResponse, HttpResponse
from django.utils import timezone
from django.apps import apps
# from django.db import IntegrityError
from .models import DailyReport, CreatorsSummaryReport, UserDepartment
from commondata.models import Department
from commondata.forms import DateForm, DateSelectionForm, DateRangeForm
from .forms import DailyReportForm
from .decorators import check_summary_report_creator, check_user_department, check_daily_user
from .utils import get_date_for_report, get_users_for_department
from shutil import copy
import os
from openpyxl import load_workbook


@login_required
@check_daily_user
def daily_reports(request):
    # Получаем текущего пользователя
    user = request.user
    # Получаем филиалы, к которым пользователь имеет отношение
    user_departments = UserDepartment.objects.filter(user=user).values_list('department', flat=True)
    # Получаем все отчёты, связанные с этими филиалами, и сортируем их по дате отчёта
    reports = DailyReport.objects.filter(department__in=user_departments).order_by('-created_at')
    # Если форма отправлена методом POST
    if request.method == 'POST':
        form = DateForm(request.POST)
        if form.is_valid():
            selected_date = form.cleaned_data['selected_date']
            # Фильтруем отчёты по выбранной дате
            reports = reports.filter(report_date=selected_date)
    else:
        form = DateForm(initial={'selected_date': get_date_for_report().strftime('%d.%m.%Y')})
    for report in reports:
        report.user_full_name = f"{report.author.last_name} {report.author.first_name}"
    summary_reports_creators = []
    first_summary_report = CreatorsSummaryReport.objects.first()
    if first_summary_report:
        summary_reports_creators = [user.username for user in first_summary_report.creators.all()]
    return render(request, 'daily/reports_list.html',
                  {'reports': reports,
                   'form': form,
                   'summary_reports_creators': summary_reports_creators})


@login_required
@check_user_department
def report_details(request, report_id):
    # Получаем объект отчёта по его id
    report = get_object_or_404(DailyReport, pk=report_id)
    # Объединяем имя и фамилию пользователя через пробел
    user_full_name = f"{report.author.last_name} {report.author.first_name}"
    return render(request, 'daily/report_details.html', {'report': report, 'user_full_name': user_full_name})


@login_required
def add_daily_report(request):
    if request.method == 'POST':
        # Создаем экземпляр формы ежедневного отчёта, передавая текущего пользователя и данные из POST-запроса
        form = DailyReportForm(request.user, request.POST)  # Передаем объект пользователя в форму
        if form.is_valid():  # Проверяем валидность данных формы
            daily_report = form.save(commit=False)  # Создаем объект ежедневного отчёта, не сохраняя его в базу данных
            daily_report.author = request.user  # Устанавливаем автора отчёта
            # # Проверяем наличие отчёта для выбранного филиала и даты
            if DailyReport.objects.filter(department=daily_report.department,
                                          report_date=daily_report.report_date).exists():
                return render(request, 'daily/denied_add_report.html',
                              {'department': daily_report.department.name,
                               'report_date': daily_report.report_date.strftime('%d.%m.%Y')})
            # try:
            #     daily_report.save()  # Сохраняем отчёт в базу данных
            # except IntegrityError:
            #     return render(request, 'daily/denied_add_report.html',
            #                   {'department': daily_report.department.name,
            #                    'report_date': daily_report.report_date.strftime('%d.%m.%Y')})
            daily_report.save()  # Сохраняем отчёт в базу данных
            return render(request, 'daily/success_page.html')
    else:
        # Получаем первое подразделение пользователя, если оно есть
        first_user_department = UserDepartment.objects.filter(user=request.user).first()
        if first_user_department:
            first_department = first_user_department.department.first()
        else:
            # Иначе получаем первое подразделение в системе
            first_department = Department.objects.first()
        # Создаем экземпляр формы ежедневного отчёта с начальными данными,
        # передавая текущего пользователя и первое доступное пользователю подразделение
        form = DailyReportForm(request.user, initial={
            'department': first_department})  # Передаем объект пользователя и начальные значения в форму
    # Отображаем шаблон add_daily_report.html с переданной формой
    return render(request, 'daily/add_daily_report.html', {'form': form})


@login_required
@check_user_department
def edit_daily_report(request, report_id):
    # Получаем отчёт по его идентификатору
    report = get_object_or_404(DailyReport, id=report_id)
    # Сохраняем дату отчёта до редактирования
    old_report_date = report.report_date
    # Получаем текущее время в часовом поясе Москвы
    current_time = timezone.now()
    # Проверяем, прошло ли меньше 1 часа с момента создания отчёта
    if current_time - report.created_at > timedelta(hours=1):
        # Если прошло более часа, переходим на страницу с сообщением об ошибке
        return render(request, 'daily/error_edit_report.html')

    if request.method == 'POST':
        # Если запрос метода POST, обрабатываем форму
        form = DailyReportForm(request.user, request.POST, instance=report)
        if form.is_valid():
            # Если форма валидна, сохраняем отчёт
            daily_report = form.save(commit=False)
            daily_report.author = request.user
            # В случае, если в отчёте изменилась дата, то проверяем наличие отчёта для выбранного филиала и даты
            if daily_report.report_date != old_report_date:
                if DailyReport.objects.filter(department=daily_report.department,
                                              report_date=daily_report.report_date).exists():
                    return render(request, 'daily/denied_add_report.html',
                                  {'department': daily_report.department.name,
                                   'report_date': daily_report.report_date.strftime('%d.%m.%Y')})
            daily_report.save()
            # Перенаправляем пользователя на страницу успешного завершения
            return render(request, 'daily/success_page.html')
    else:
        # Если запрос метода GET, отображаем форму для редактирования
        form = DailyReportForm(request.user, instance=report)
    # Возвращаем HTML-страницу с формой для редактирования отчёта
    return render(request, 'daily/edit_daily_report.html', {'form': form, 'report_id': report_id})


@login_required
@check_summary_report_creator
def summary_report(request):
    date = None
    if request.method == 'POST':
        # Если запрос метода POST, обрабатываем форму
        form = DateSelectionForm(request.POST)
        if form.is_valid():
            # Если форма действительна, извлекаем выбранную дату из формы
            date = form.cleaned_data['report_date']
    else:
        form = DateSelectionForm(initial={'report_date': get_date_for_report().strftime('%d.%m.%Y')})
        date = get_date_for_report()
    reports = DailyReport.objects.filter(report_date=date)
    # Получаем список подразделений и пользователей, без отчётов за дату date
    departments_without_reports = [': '.join([department.name, ', '.join(get_users_for_department(department.name))])
                                   for department in Department.objects.all().exclude(dailyreport__report_date=date)
                                   .order_by('name')]
    return render(request, 'daily/summary_report.html',
                  {'form': form, 'reports': reports,
                   'departments_without_reports': departments_without_reports})


@login_required
@check_summary_report_creator
def generate_summary_report(request):
    """
    Генерация сводного ежедневного отчёта
    """
    if request.method == 'POST':
        # Получаем данные формы
        form = DateSelectionForm(request.POST)
        if form.is_valid():
            # Извлекаем выбранную дату из формы
            selected_date = form.cleaned_data['report_date']
            # Получаем экземпляр класса конфигурации приложения daily
            daily_config = apps.get_app_config('daily')
            # Получаем путь к папке для сохранения отчётов
            path_to_reports = daily_config.PATH_TO_SAVE
            # Префикс названия отчёта
            prefix_report_name = 'Ежедневный отчёт по охране'
            # Формируем имя файла отчёта на основе выбранной даты
            report_name = os.path.join(path_to_reports, f'{prefix_report_name} за '
                                                        f'{selected_date.strftime("%d.%m.%Y")}.xlsx')
            # Копируем шаблон отчёта
            copy(os.path.join(path_to_reports, 'daily_summary_report_blank.xlsx'), report_name)
            # Получаем все отчёты за выбранную дату, отсортированные по дате
            reports = DailyReport.objects.filter(report_date=selected_date).order_by('-report_date')
            # Загружаем созданный отчёт в Excel
            report_workbook = load_workbook(report_name)
            report_sheet = report_workbook.active
            # Словарь, соотносящий названия подразделений с соответствующими столбцами в Excel
            departments_cols = {k: v for k, v in zip(('Марий Эл и Чувашии', 'Ульяновский', 'Удмуртский', 'Свердловский',
                                                      'Саратовский', 'Самарский', 'Пермский', 'Оренбургский',
                                                      'Нижегородский', 'Мордовский', 'Коми', 'Кировский',
                                                      'Владимирский', 'Пензенский'), 'CDEFGHIJKLMNOP')}
            # Копируем данные из отчётов филиалов в сводный отчёт
            for report in reports:
                for idx, line in enumerate(tuple(range(3, 12)) + (16, 17), start=1):
                    report_sheet[f'{departments_cols[report.department.name]}{line}'] = getattr(report,
                                                                                                f'field_{idx}')
            # Сохраняем изменения в отчёте
            report_workbook.save(report_name)
            # Возвращаем файл отчёта в HTTP-ответе
            response = FileResponse(open(report_name, 'rb'), as_attachment=True,
                                    filename=report_name)
            return response


@login_required
def success_page(request):
    return render(request, 'daily/success_page.html')


@login_required
@check_summary_report_creator
def summary_weekly_report(request):

    def get_week_range(input_date=None):
        """
        Возвращает даты понедельника и воскресенья недели, к которой принадлежит указанная дата.
        Если дата не указана, используется сегодняшняя дата.
        Аргументы:
        input_date (datetime.date или datetime.datetime, опционально):
        Дата, для которой необходимо определить диапазон недели.
        Если не указана, используется текущая дата.
        Возвращает:
        tuple: Две даты (datetime.date), представляющие понедельник и воскресенье соответствующей недели.
        """
        if input_date is None:
            input_date = datetime.today().date()
        else:
            if isinstance(input_date, datetime):
                input_date = input_date.date()
        # Определяем день недели: понедельник - 0, ..., воскресенье - 6
        weekday = input_date.weekday()
        # Определяем дату понедельника и воскресенья
        start_date = input_date - timedelta(days=weekday)
        end_date = start_date + timedelta(days=6)
        return start_date, end_date
    start_date, end_date = get_week_range()
    reports = DailyReport.objects.filter(report_date__range=(start_date, end_date)).order_by('-report_date')
    for report in reports:
        report.user_full_name = f"{report.author.last_name} {report.author.first_name}"
    if request.method == "POST":
        form = DateRangeForm(request.POST)
        if form.is_valid():
            # Обработка данных формы
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
            reports = DailyReport.objects.filter(report_date__range=(start_date, end_date)).order_by('-report_date')
            for report in reports:
                report.user_full_name = f"{report.author.last_name} {report.author.first_name}"
    else:
        print(type(start_date))
        form = DateRangeForm(initial={'start_date': start_date.strftime('%d.%m.%Y'),
                                      'end_date': end_date.strftime('%d.%m.%Y')})
    return render(request, 'daily/date_range.html', {'reports': reports, 'form': form})


def generate_weekly_summary_report(request):
    if request.method == "POST":
        form = DateRangeForm(request.POST)
        if form.is_valid():
            # Извлекаем выбранные даты из формы
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
            # Получаем экземпляр класса конфигурации приложения daily
            daily_config = apps.get_app_config('daily')
            # Получаем путь к папке для сохранения отчётов
            path_to_reports = daily_config.PATH_TO_SAVE
            # Префикс названия отчёта
            prefix_report_name = 'Сводный отчёт по охране'
            # Формируем имя файла отчёта на основе выбранной даты
            report_name = os.path.join(path_to_reports, f'{prefix_report_name} с '
                                                        f'{start_date.strftime("%d.%m.%Y")} по '
                                                        f'{end_date.strftime("%d.%m.%Y")}.xlsx')
            # Копируем шаблон отчёта
            copy(os.path.join(path_to_reports, 'daily_summary_report_blank.xlsx'), report_name)
            # # Получаем все отчёты за выбранные даты
            # reports = DailyReport.objects.filter(report_date__range=(start_date, end_date))
            # Загружаем созданный отчёт в Excel
            report_workbook = load_workbook(report_name)
            report_sheet = report_workbook.active
            # Словарь, соотносящий названия подразделений с соответствующими столбцами в Excel
            departments_cols = {k: v for k, v in zip(('Марий Эл и Чувашии', 'Ульяновский', 'Удмуртский', 'Свердловский',
                                                      'Саратовский', 'Самарский', 'Пермский', 'Оренбургский',
                                                      'Нижегородский', 'Мордовский', 'Коми', 'Кировский',
                                                      'Владимирский', 'Пензенский'), 'CDEFGHIJKLMNOP')}
            weekly_sums = {}
            # Считаем суммы за указанный диапазон дат
            for department in Department.objects.all():
                weekly_sums[department.name] = DailyReport.get_weekly_sums(department, start_date, end_date)
            # Копируем посчитанные суммы в сводный отчёт
            for department in Department.objects.all():
                for idx, line in enumerate(tuple(range(3, 12)) + (16, 17), start=1):
                    report_sheet[f'{departments_cols[department.name]}{line}'] = (
                        0 if weekly_sums[department.name].get(f'total_field_{idx}') is None
                        else weekly_sums[department.name][f'total_field_{idx}']
                    )
            report_sheet['A16'] = 'Количество проведенных проверок СБ'
            report_sheet['A17'] = 'Количество направленных претензионных писем'
            # Высота строки
            report_sheet.row_dimensions[16].height = 15
            report_sheet.row_dimensions[17].height = 15
            # Сохраняем изменения в отчёте
            report_workbook.save(report_name)
            # Возвращаем файл отчёта в HTTP-ответе
            response = FileResponse(open(report_name, 'rb'), as_attachment=True,
                                    filename=report_name)
            return response
        else:
            return HttpResponse('Invalid form data')  # Сообщение о неверных данных формы
    else:
        return HttpResponse('Only POST requests are allowed')  # Сообщение о неправильном методе запроса
