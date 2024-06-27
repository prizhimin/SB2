from django.contrib.auth.decorators import login_required
from django.http import FileResponse, HttpResponse
from django.shortcuts import redirect
from django.shortcuts import render, get_object_or_404
from django.apps import apps
from .forms import ReportForm
from .models import (SemiAnnual2024Report, SemiAnnual2024CreatorsSummaryReport, SemiAnnual2024UserCompany,
                     SemiAnnual2024Company)
from shutil import copy
from os import path
from openpyxl import load_workbook

@login_required
def create_report(request):
    if request.method == 'POST':
        form = ReportForm(request.POST)
        if form.is_valid():
            report = form.save(commit=False)
            report.user = request.user  # Установите поле user
            report.save()
            return redirect('report_list')
    else:
        form = ReportForm()
    return render(request, 'sixmonths2024/report_form.html', {'form': form})


@login_required
def update_report(request, pk):
    report = get_object_or_404(SemiAnnual2024Report, pk=pk)
    if request.method == 'POST':
        form = ReportForm(request.POST, instance=report)
        if form.is_valid():
            form.save()
            return redirect('report_list')
    else:
        form = ReportForm(instance=report)
    return render(request, 'sixmonths2024/report_form.html', {'form': form})


@login_required
def delete_report(request, pk):
    report = get_object_or_404(SemiAnnual2024Report, pk=pk)

    if request.method == 'POST':
        report.delete()
        # messages.success(request, 'Отчет успешно удален.')
        return redirect('report_list')

    return render(request, 'sixmonths2024/report_confirm_delete.html', {'report': report})


@login_required
def report_list(request):
    user = request.user

    creators_summary_report = SemiAnnual2024CreatorsSummaryReport.objects.first()
    if creators_summary_report and user in creators_summary_report.creators.all():
        reports = SemiAnnual2024Report.objects.all().order_by('company')
        return render(request, 'sixmonths2024/report_list.html', {'reports': reports})
    else:
        user_company = SemiAnnual2024UserCompany.objects.get(user=user)
        company = user_company.companies.first()

        # Если для этой компании уже есть отчет, получаем его
        report = SemiAnnual2024Report.objects.filter(company=company).first()

        if request.method == 'POST':
            form = ReportForm(request.POST, instance=report)
            if form.is_valid():
                report = form.save(commit=False)
                report.user = request.user  # Устанавливаем текущего пользователя
                report.save()
                return redirect('report_list')
        else:
            form = ReportForm(instance=report, initial={'company': company})
            user_company = SemiAnnual2024UserCompany.objects.filter(user=user).first()
            available_companies = user_company.companies.all()  # Компании, доступные пользователю
            form.fields['company'].queryset = available_companies  # Ограничиваем доступные компании
        return render(request, 'sixmonths2024/report_form.html', {'form': form})


def generate_sixmonths_2024_summary_report(request):
    # Получаем экземпляр класса конфигурации приложения sixmonths2024
    sixmonths2024_config = apps.get_app_config('sixmonths2024')
    # Получаем путь к папке для сохранения отчётов
    path_to_reports = sixmonths2024_config.PATH_TO_SAVE
    # Префикс названия отчёта
    report_name = 'Сводный отчёт за 6 месяцев 2024 г..xlsx'
    # Формируем имя файла отчёта на основе выбранной даты
    report_name = path.join(path_to_reports, f'{report_name}')
    # Копируем шаблон отчёта
    copy(path.join(path_to_reports, 'sixmonths2024_blank.xlsx'), report_name)
    # Загружаем созданный отчёт в Excel
    workbook = load_workbook(report_name)
    # Получаем список названий всех компаний из модели SemiAnnual2024Company
    company_names = SemiAnnual2024Company.objects.values_list('name', flat=True)
    # Итерируемся по списку названий компаний
    for company_name in company_names:
        # Получаем объект компании по её названию
        company = SemiAnnual2024Company.objects.get(name=company_name)
        # Проверяем, существует ли отчет для этой компании
        if SemiAnnual2024Report.objects.filter(company=company).exists():
            # Объект отчёт филиал company_name
            report = SemiAnnual2024Report.objects.filter(company=company).first()
            # Получаем номер строки в таблице
            line_number = company.line_number
            print(f'Номер строки в таблице {line_number}')
            # Начинаем формировать отчёт
            for idx_field in range(1, 52):
                match idx_field:
                    # закладка кадры поля 1,2
                    case 1 | 2:
                        cols = 'CD'
                        sheet = workbook[workbook.sheetnames[0]]
                        sheet[f'{cols[idx_field - 1]}{line_number-1}'] = getattr(report, f'field_{idx_field}')
                    # закладка уд и дз поля 3..19
                    case n if 3 <= n <= 19:
                        cols = 'CDEFGHIJKMNOPQRST'
                        sheet = workbook[workbook.sheetnames[1]]
                        sheet[f'{cols[idx_field - 3]}{line_number}'] = getattr(report, f'field_{idx_field}')
                        # print(f'Колонка {cols[idx_field - 3]}')
                    # закладка антикорпоратив и коррупционные поля 20..25
                    case n if 20 <= n <= 25:
                        sheet = workbook[workbook.sheetnames[2]]
                        cols = 'CDEFGH'
                        sheet[f'{cols[idx_field - 20]}{line_number-1}'] = getattr(report, f'field_{idx_field}')
                    # закладка охрана поля 26..33
                    case n if 26 <= n <= 33:
                        sheet = workbook[workbook.sheetnames[3]]
                        cols = 'CDEFHIJK'
                        sheet[f'{cols[idx_field - 26]}{line_number}'] = getattr(report, f'field_{idx_field}')
                    # закладка проверка физ. юр. лиц поля 34..37
                    case n if 34 <= n <= 37:
                        sheet = workbook[workbook.sheetnames[4]]
                        cols = 'CDEF'
                        sheet[f'{cols[idx_field - 34]}{line_number}'] = getattr(report, f'field_{idx_field}')
                    # закладка травматизм поля 38..43
                    case n if 38 <= n <= 43:
                        sheet = workbook[workbook.sheetnames[5]]
                        cols = 'CDEGHIJ'
                        sheet[f'{cols[idx_field - 38]}{line_number}'] = getattr(report, f'field_{idx_field}')
                    # закладка аварийность поля 44..47
                    case n if 44 <= n <= 47:
                        sheet = workbook[workbook.sheetnames[6]]
                        cols = 'CDEFGH'
                        sheet[f'{cols[idx_field - 44]}{line_number}'] = getattr(report, f'field_{idx_field}')
                    # закладка аналитика поля 48..51
                    case n if 48 <= n <= 51:
                        sheet = workbook[workbook.sheetnames[7]]
                        cols = 'CDEF'
                        sheet[f'{cols[idx_field - 48]}{line_number}'] = getattr(report, f'field_{idx_field}')
    workbook.save(report_name)
    # Возвращаем файл отчёта в HTTP-ответе
    return FileResponse(open(report_name, 'rb'), as_attachment=True, filename=report_name)
